from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha1
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent

MONTH_SHEETS = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

TEXT_MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

TRANSACTION_RE = re.compile(
    r"^(?P<date>\d{2}/\d{2})\s+(?P<description>.+?)\s+R\$\s+(?P<amount>-?[\d\.,]+)$"
)


@dataclass
class Transaction:
    date: datetime
    description: str
    amount: Decimal
    bank_heading: str
    target_category: str
    target_type: str
    source_pdf: str
    fingerprint: str


def load_config(config_path: Path) -> dict:
    with config_path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def open_workbook(workbook_path: Path, data_only: bool = False):
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    return load_workbook(workbook_path, keep_vba=keep_vba, data_only=data_only)


def load_categories(categories_path: Path | None = None) -> dict:
    resolved_path = categories_path or (BASE_DIR / "categories.json")
    with resolved_path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_learned_categories(learned_path: Path | None = None) -> dict[str, str]:
    resolved_path = learned_path or (BASE_DIR / "learned_categories.json")
    if not resolved_path.exists():
        return {}

    with resolved_path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)

    return {
        normalize_description(key): value
        for key, value in payload.get("description_to_category", {}).items()
    }


def save_learned_categories(description_map: dict[str, str], learned_path: Path | None = None) -> None:
    resolved_path = learned_path or (BASE_DIR / "learned_categories.json")
    payload = {"description_to_category": dict(sorted(description_map.items()))}
    with resolved_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def normalize_description(value: str) -> str:
    cleaned = normalize_spaces(value).upper()
    return cleaned.replace(" REFEICTOLEDO", " REFEICAO TOLEDO").replace(" REFEICO", " REFEICAO")


def parse_decimal_br(value: str) -> Decimal:
    try:
        return Decimal(value.replace(".", "").replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"Valor inválido: {value}") from exc


def infer_statement_year(statement_close_date: datetime, transaction_month: int) -> int:
    if transaction_month > statement_close_date.month:
        return statement_close_date.year - 1
    return statement_close_date.year


def extract_statement_close_date(full_text: str) -> datetime:
    match = re.search(r"Fatura fechada em (\d{2}/\d{2}/\d{4})", full_text)
    if match:
        return datetime.strptime(match.group(1), "%d/%m/%Y")

    month_match = re.search(r"esta é sua fatura de\s+([A-Za-zçÇãõáéíóú]+)", full_text, re.IGNORECASE)
    if month_match:
        month_name = month_match.group(1).strip().lower()
        month_number = TEXT_MONTHS.get(month_name)
        if month_number:
            today = datetime.today()
            return datetime(today.year, month_number, 1)

    raise ValueError("Não foi possível identificar a data de fechamento da fatura.")


def parse_pdf_transactions(pdf_path: Path, category_rules: dict) -> list[tuple[str, str, Decimal]]:
    reader = PdfReader(str(pdf_path))
    page_texts = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n".join(page_texts)
    close_date = extract_statement_close_date(full_text)
    heading_to_category = category_rules["heading_to_category"]

    transactions: list[tuple[str, str, Decimal]] = []
    current_heading = ""
    seen_section = False

    for raw_line in full_text.splitlines():
        line = normalize_spaces(raw_line)
        if not line:
            continue
        if "Lançamentos nesta fatura" in line:
            seen_section = True
            continue
        if not seen_section:
            continue
        if line.startswith("Total da Fatura"):
            break
        if line.startswith("Página "):
            continue
        if "(Cartão " in line or line.startswith("Data Descrição País Valor"):
            continue
        if line in heading_to_category:
            current_heading = line
            continue
        match = TRANSACTION_RE.match(line)
        if not match:
            continue

        day_month = match.group("date")
        description = normalize_description(match.group("description"))
        amount = parse_decimal_br(match.group("amount"))
        tx_month = int(day_month.split("/")[1])
        tx_year = infer_statement_year(close_date, tx_month)
        tx_date = datetime.strptime(f"{day_month}/{tx_year}", "%d/%m/%Y")
        transactions.append((current_heading, description, amount))
        transactions[-1] = (current_heading, tx_date.strftime("%Y-%m-%d"), description, amount)

    parsed: list[tuple[str, str, Decimal]] = []
    for heading, iso_date, description, amount in transactions:
        parsed.append((heading, f"{iso_date}|{description}", amount))
    return parsed


def load_category_index(workbook_path: Path) -> dict[str, str]:
    wb = open_workbook(workbook_path, data_only=False)
    ws = wb["Categorias"]
    category_index: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        essential = ws.cell(row, 2).value
        non_essential = ws.cell(row, 3).value
        if essential:
            category_index[str(essential)] = "essencial"
        if non_essential:
            category_index[str(non_essential)] = "nao_essencial"
    wb.close()
    return category_index


def choose_category(
    description: str,
    bank_heading: str,
    category_index: dict[str, str],
    category_rules: dict,
    learned_categories: dict[str, str],
) -> str:
    learned_category = learned_categories.get(description)
    if learned_category in category_index:
        return learned_category

    for rule in category_rules["merchant_rules"]:
        pattern = rule["pattern"]
        category = rule["category"]
        if re.search(pattern, description, re.IGNORECASE):
            if category in category_index:
                return category

    fallback = category_rules["heading_to_category"].get(bank_heading)
    if fallback in category_index:
        return fallback

    if "💳 Cartão / Financeiro" in category_index:
        return "💳 Cartão / Financeiro"
    raise ValueError("Nenhuma categoria válida encontrada na planilha.")


def build_transactions(
    pdf_path: Path,
    workbook_path: Path,
    categories_path: Path | None = None,
    learned_path: Path | None = None,
) -> list[Transaction]:
    category_rules = load_categories(categories_path)
    category_index = load_category_index(workbook_path)
    learned_categories = load_learned_categories(learned_path)
    parsed_rows = parse_pdf_transactions(pdf_path, category_rules)
    transactions: list[Transaction] = []

    for bank_heading, packed_payload, amount in parsed_rows:
        iso_date, description = packed_payload.split("|", 1)
        date = datetime.strptime(iso_date, "%Y-%m-%d")
        category = choose_category(description, bank_heading, category_index, category_rules, learned_categories)
        target_type = category_index[category]
        if amount <= 0:
            continue

        fingerprint_base = f"{pdf_path.name}|{iso_date}|{description}|{amount:.2f}"
        transactions.append(
            Transaction(
                date=date,
                description=description,
                amount=amount,
                bank_heading=bank_heading,
                target_category=category,
                target_type=target_type,
                source_pdf=pdf_path.name,
                fingerprint=sha1(fingerprint_base.encode("utf-8")).hexdigest(),
            )
        )

    return transactions


def ensure_import_sheet(workbook) -> object:
    sheet_name = "_Importacoes"
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    ws = workbook.create_sheet(sheet_name)
    ws.sheet_state = "hidden"
    headers = [
        "imported_at",
        "source_pdf",
        "fingerprint",
        "transaction_date",
        "description",
        "amount",
        "bank_heading",
        "target_category",
        "target_type",
        "month_sheet",
        "row_number",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    return ws


def load_existing_fingerprints(import_ws) -> set[str]:
    fingerprints: set[str] = set()
    for row in range(2, import_ws.max_row + 1):
        value = import_ws.cell(row, 3).value
        if value:
            fingerprints.add(str(value))
    return fingerprints


def next_empty_row(sheet, columns: tuple[int, int, int], start_row: int = 7) -> int:
    date_col, value_col, category_col = columns
    for row in range(start_row, 10000):
        values = [sheet.cell(row, col).value for col in (date_col, value_col, category_col)]
        if all(value in (None, "") for value in values):
            return row
    raise ValueError(f"Não foi encontrado espaço livre na aba {sheet.title}.")


def column_triplet(target_type: str) -> tuple[int, int, int]:
    if target_type == "essencial":
        return 5, 6, 7
    if target_type == "nao_essencial":
        return 9, 10, 11
    raise ValueError(f"Tipo de gasto desconhecido: {target_type}")


def sync_learning_sources(
    workbook,
    import_ws,
    category_index: dict[str, str],
    learned_path: Path | None = None,
) -> dict[str, str]:
    learned_categories = load_learned_categories(learned_path)
    changed = False

    for row in range(2, import_ws.max_row + 1):
        description = import_ws.cell(row, 5).value
        stored_category = import_ws.cell(row, 8).value
        stored_type = import_ws.cell(row, 9).value
        month_sheet_name = import_ws.cell(row, 10).value
        row_number = import_ws.cell(row, 11).value

        if not description:
            continue

        normalized_description = normalize_description(str(description))

        if stored_category in category_index and learned_categories.get(normalized_description) != stored_category:
            learned_categories[normalized_description] = str(stored_category)
            changed = True

        if not month_sheet_name or not row_number or month_sheet_name not in workbook.sheetnames:
            continue

        month_ws = workbook[month_sheet_name]
        columns = column_triplet(str(stored_type))
        current_category = month_ws.cell(int(row_number), columns[2]).value
        if not current_category:
            continue

        current_category = str(current_category)
        if current_category not in category_index:
            continue

        if current_category != stored_category:
            import_ws.cell(row, 8).value = current_category
            import_ws.cell(row, 9).value = category_index[current_category]
            learned_categories[normalized_description] = current_category
            changed = True

    if changed:
        save_learned_categories(learned_categories, learned_path)

    return learned_categories


def write_transactions(
    workbook_path: Path,
    transactions: Iterable[Transaction],
    output_path: Path | None = None,
    learned_path: Path | None = None,
) -> dict[str, int]:
    wb = open_workbook(workbook_path, data_only=False)
    import_ws = ensure_import_sheet(wb)
    category_index = load_category_index(workbook_path)
    learned_categories = sync_learning_sources(wb, import_ws, category_index, learned_path=learned_path)
    existing_fingerprints = load_existing_fingerprints(import_ws)
    imported_count = 0
    skipped_count = 0

    for tx in sorted(transactions, key=lambda item: item.date):
        if tx.fingerprint in existing_fingerprints:
            skipped_count += 1
            continue

        month_sheet_name = MONTH_SHEETS[tx.date.month]
        month_ws = wb[month_sheet_name]
        columns = column_triplet(tx.target_type)
        row = next_empty_row(month_ws, columns)

        month_ws.cell(row, columns[0]).value = tx.date
        month_ws.cell(row, columns[1]).value = float(tx.amount)
        month_ws.cell(row, columns[2]).value = tx.target_category

        import_row = import_ws.max_row + 1
        import_ws.cell(import_row, 1).value = datetime.now()
        import_ws.cell(import_row, 2).value = tx.source_pdf
        import_ws.cell(import_row, 3).value = tx.fingerprint
        import_ws.cell(import_row, 4).value = tx.date
        import_ws.cell(import_row, 5).value = tx.description
        import_ws.cell(import_row, 6).value = float(tx.amount)
        import_ws.cell(import_row, 7).value = tx.bank_heading
        import_ws.cell(import_row, 8).value = tx.target_category
        import_ws.cell(import_row, 9).value = tx.target_type
        import_ws.cell(import_row, 10).value = month_sheet_name
        import_ws.cell(import_row, 11).value = row
        existing_fingerprints.add(tx.fingerprint)
        learned_categories[tx.description] = tx.target_category
        imported_count += 1

    final_path = output_path or workbook_path
    wb.save(final_path)
    wb.close()
    save_learned_categories(learned_categories, learned_path)
    return {"imported": imported_count, "skipped": skipped_count}


def process_invoice(
    pdf_path: Path,
    workbook_path: Path,
    output_path: Path | None,
    dry_run: bool,
    categories_path: Path | None = None,
    learned_path: Path | None = None,
) -> dict[str, int]:
    transactions = build_transactions(
        pdf_path,
        workbook_path,
        categories_path=categories_path,
        learned_path=learned_path,
    )
    if dry_run:
        print(f"Arquivo: {pdf_path.name}")
        for tx in transactions:
            print(
                f"{tx.date:%Y-%m-%d} | R$ {tx.amount:.2f} | "
                f"{tx.target_category} | {tx.description}"
            )
        return {"imported": 0, "skipped": 0, "found": len(transactions)}

    result = write_transactions(
        workbook_path,
        transactions,
        output_path=output_path,
        learned_path=learned_path,
    )
    result["found"] = len(transactions)
    return result


def watch_folder(config: dict) -> None:
    invoice_dir = Path(config["invoice_dir"])
    workbook_path = Path(config["workbook_path"])
    processed_dir = Path(config["processed_dir"])
    output_path = Path(config["output_workbook_path"]) if config.get("output_workbook_path") else None
    categories_path = Path(config["categories_path"]) if config.get("categories_path") else None
    learned_path = Path(config["learned_categories_path"]) if config.get("learned_categories_path") else None
    poll_seconds = int(config.get("poll_seconds", 15))

    processed_dir.mkdir(parents=True, exist_ok=True)
    print(f"Monitorando {invoice_dir} a cada {poll_seconds}s...")

    while True:
        pdf_files = sorted(invoice_dir.glob("*.pdf"))
        for pdf_path in pdf_files:
            archived_path = processed_dir / pdf_path.name
            if archived_path.exists():
                continue

            try:
                result = process_invoice(
                    pdf_path,
                    workbook_path,
                    output_path=output_path,
                    dry_run=False,
                    categories_path=categories_path,
                    learned_path=learned_path,
                )
                pdf_path.replace(archived_path)
                print(
                    f"[OK] {pdf_path.name}: encontradas {result['found']}, "
                    f"importadas {result['imported']}, ignoradas {result['skipped']}."
                )
            except Exception as exc:
                print(f"[ERRO] {pdf_path.name}: {exc}", file=sys.stderr)

        time.sleep(poll_seconds)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Importa gastos de faturas PDF para a planilha de controle financeiro."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    import_parser = subparsers.add_parser("import-pdf", help="Processa uma única fatura PDF.")
    import_parser.add_argument("--pdf", required=True, help="Caminho do PDF da fatura.")
    import_parser.add_argument("--workbook", required=True, help="Caminho da planilha XLSM.")
    import_parser.add_argument(
        "--output-workbook",
        help="Opcional: caminho de saída para salvar uma cópia atualizada da planilha.",
    )
    import_parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Mostra os lançamentos interpretados sem gravar nada.",
    )
    import_parser.add_argument(
        "--categories",
        help="Opcional: caminho do arquivo JSON com regras de categorias.",
    )
    import_parser.add_argument(
        "--learned-categories",
        help="Opcional: caminho do arquivo JSON com categorias aprendidas.",
    )

    watch_parser = subparsers.add_parser("watch", help="Monitora a pasta configurada de faturas.")
    watch_parser.add_argument(
        "--config",
        default="config.json",
        help="Caminho do arquivo JSON de configuração.",
    )

    return parser


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    parser = build_parser()
    args = parser.parse_args()

    if args.command == "import-pdf":
        output_path = Path(args.output_workbook) if args.output_workbook else None
        categories_path = Path(args.categories) if args.categories else None
        learned_path = Path(args.learned_categories) if args.learned_categories else None
        result = process_invoice(
            pdf_path=Path(args.pdf),
            workbook_path=Path(args.workbook),
            output_path=output_path,
            dry_run=args.dry_run,
            categories_path=categories_path,
            learned_path=learned_path,
        )
        print(json.dumps(result, ensure_ascii=False))
        return 0

    if args.command == "watch":
        config = load_config(Path(args.config))
        watch_folder(config)
        return 0

    parser.print_help()
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
